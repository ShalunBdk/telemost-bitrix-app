# -*- coding: utf-8 -*-
"""
Flask веб-приложение для управления FAQ и переобучения ChromaDB
"""

from flask import Flask, Blueprint, render_template, request, jsonify, redirect, url_for, make_response
from flask_cors import CORS
from werkzeug.middleware.proxy_fix import ProxyFix
import uuid
import sys
import logging
import os
import signal
import requests
import jwt
import re
from io import BytesIO, TextIOWrapper
import csv
from dotenv import load_dotenv
from datetime import datetime
from urllib.parse import quote

# Библиотеки для экспорта
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Добавляем корневую директорию проекта в PYTHONPATH
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from src.core import database
from src.core import logging_config
from src.web.middleware import get_allowed_origins, is_production, cors_origin_validator, require_bitrix24_auth
from src.web.bitrix24_integration import handle_install, handle_index, handle_app
from src.web.bitrix24_permissions import bitrix24_permissions_bp

# Загружаем переменные окружения
load_dotenv()

os.environ["ANONYMIZED_TELEMETRY"] = "False"

import chromadb
from chromadb.utils import embedding_functions

# Определяем пути к статическим файлам и шаблонам
current_dir = os.path.dirname(os.path.abspath(__file__))
static_folder = os.path.join(current_dir, 'static')
template_folder = os.path.join(current_dir, 'templates')
BASE_PATH = os.getenv('BASE_PATH', '').rstrip('/')

# Создаём Flask приложение
app = Flask(__name__,
            static_folder=static_folder,
            template_folder=template_folder,
            static_url_path=f"{BASE_PATH}/static")

# ProxyFix middleware - правильная обработка X-Script-Name от nginx
# Это позволяет Flask корректно генерировать URL при работе за reverse proxy
app.wsgi_app = ProxyFix(
    app.wsgi_app,
    x_for=1,       # X-Forwarded-For
    x_proto=1,     # X-Forwarded-Proto
    x_host=1,      # X-Forwarded-Host
    x_prefix=1     # X-Script-Name → SCRIPT_NAME (BASE_PATH)
)

app.config['JSON_AS_ASCII'] = False
app.config['BASE_PATH'] = BASE_PATH  # Для использования в templates
app.config['PREFERRED_URL_SCHEME'] = 'https'  # Для генерации HTTPS URL

# Настройка CORS для работы с Битрикс24
# Получаем список разрешённых origins
allowed_origins = get_allowed_origins()

# Если список пустой, добавляем wildcard для development
if not allowed_origins or not is_production():
    # В development разрешаем все origins
    allowed_origins = ['*']

CORS(app,
     origins=allowed_origins,
     supports_credentials=True,
     allow_headers=['Content-Type', 'Authorization', 'X-Requested-With'],
     methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS']
)

logging_config.configure_root_logger(level=logging.INFO)
logging.getLogger('werkzeug').setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# Security headers для работы в iframe Битрикс24
@app.after_request
def set_security_headers(response):
    """Устанавливаем security headers для работы в iframe"""

    # Content Security Policy для iframe
    if is_production():
        # В production строго ограничиваем
        bitrix_domain = os.getenv('BITRIX24_DOMAIN', '')
        if bitrix_domain:
            response.headers['Content-Security-Policy'] = (
                f"frame-ancestors 'self' https://{bitrix_domain} https://*.bitrix24.ru https://*.bitrix24.com;"
                f"script-src 'self' 'unsafe-inline' https://cdn.quilljs.com https://cdn.jsdelivr.net/ https://api.bitrix24.com/; "
                f"style-src 'self' 'unsafe-inline' https://cdn.quilljs.com https://cdn.jsdelivr.net/; "
                f"font-src 'self' data:;"
            )
    else:
        # В development разрешаем все для тестирования
        response.headers['Content-Security-Policy'] = (
            "frame-ancestors *; "
            "script-src 'self' 'unsafe-inline' https://cdn.quilljs.com https://cdn.jsdelivr.net/ https://api.bitrix24.com/; "
            "style-src 'self' 'unsafe-inline' https://cdn.quilljs.com https://cdn.jsdelivr.net/; "
            "font-src 'self' data:;"
        )

    return response

# Конфигурация
MODEL_NAME = os.getenv("MODEL_NAME", "paraphrase-multilingual-MiniLM-L12-v2")

# Эндпоинты для уведомления ботов (поддержка Docker и localhost)
# В Docker используем имена контейнеров из переменных окружения
TELEGRAM_BOT_HOST = os.getenv('TELEGRAM_BOT_HOST', '127.0.0.1')
BITRIX24_BOT_HOST = os.getenv('BITRIX24_BOT_HOST', '127.0.0.1')

TELEGRAM_BOT_RELOAD_URL = f"http://{TELEGRAM_BOT_HOST}:5001/reload"
TELEGRAM_BOT_RELOAD_SETTINGS_URL = f"http://{TELEGRAM_BOT_HOST}:5001/reload-settings"

BITRIX24_BOT_RELOAD_URL = f"http://{BITRIX24_BOT_HOST}:5002/api/reload-chromadb"
BITRIX24_BOT_RELOAD_SETTINGS_URL = f"http://{BITRIX24_BOT_HOST}:5002/api/reload-settings"

# Список всех ботов для уведомления
ALL_BOT_RELOAD_URLS = [TELEGRAM_BOT_RELOAD_URL, BITRIX24_BOT_RELOAD_URL]
ALL_BOT_RELOAD_SETTINGS_URLS = [TELEGRAM_BOT_RELOAD_SETTINGS_URL, BITRIX24_BOT_RELOAD_SETTINGS_URL]

# Инициализация ChromaDB (поддержка Docker путей)
CHROMA_PATH = os.getenv('CHROMA_PATH', './data/chroma_db')
chroma_client = chromadb.PersistentClient(path=CHROMA_PATH)
embedding_func = embedding_functions.SentenceTransformerEmbeddingFunction(model_name=MODEL_NAME)

# Создаем Blueprint для админ-панели
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


# Защита всех роутов админки в production режиме
@admin_bp.before_request
def check_admin_access():
    """
    Проверка доступа ко всем роутам админки
    В production режиме требует:
    - Запрос с разрешенного Origin (Bitrix24) - для всех запросов
    - JWT токен - только для API запросов (POST/PUT/DELETE и /api/*)

    HTML страницы (GET /admin/, /admin/logs) доступны из Bitrix24 без токена
    для первоначальной загрузки iframe и OAuth авторизации
    """
    if not is_production():
        return  # В dev режиме нет проверок

    from src.web.middleware import check_cors_origin

    # 1. Проверяем Origin для ВСЕХ запросов
    origin = request.headers.get('Origin')
    if not origin:
        # Пробуем получить из Referer
        referer = request.headers.get('Referer', '')
        if referer:
            match = re.match(r'^(https?://[^/]+)', referer)
            if match:
                origin = match.group(1)

    if not origin or not check_cors_origin(origin):
        return jsonify({
            'error': 'Доступ запрещен',
            'message': 'Доступ к админ-панели возможен только через Битрикс24'
        }), 403

    # 2. Проверяем JWT токен только для API запросов
    # GET запросы HTML страниц разрешены (для загрузки iframe и OAuth)
    # В Blueprint request.path уже без префикса /admin, поэтому проверяем /api/
    is_api_request = (
        request.method in ['POST', 'PUT', 'DELETE'] or  # Любые изменяющие запросы
        request.path.startswith('/api/') or  # API эндпоинты в Blueprint
        '/api/' in request.path  # Дополнительная проверка
    )

    if is_api_request:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Требуется авторизация'}), 401

        token = auth_header.replace('Bearer ', '')
        try:
            JWT_SECRET = os.getenv('JWT_SECRET', 'supersecretkey')
            payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])

            # Добавляем данные пользователя в request
            request.user_id = payload.get('id')
            request.user_role = payload.get('role')
            request.username = payload.get('username')
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Токен истек'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Неверный токен'}), 401


def retrain_chromadb():
    """
    Переобучение ChromaDB на основе данных из базы
    """
    try:
        # Удаляем старую коллекцию
        try:
            chroma_client.delete_collection(name="faq_collection")
            logger.info("Старая коллекция удалена")
        except Exception as e:
            logger.info(f"Коллекции не было или ошибка удаления: {e}")

        # Создаем новую коллекцию
        collection = chroma_client.create_collection(
            name="faq_collection",
            embedding_function=embedding_func,
            metadata={"hnsw:space": "cosine"}
        )
        logger.info("Создана новая коллекция")

        # Получаем все FAQ из базы
        all_faqs = database.get_all_faqs()
        if not all_faqs:
            logger.warning("В базе нет данных для обучения")
            return {"success": False, "message": "В базе нет данных"}

        # Подготавливаем данные для ChromaDB
        documents, metadatas, ids = [], [], []
        for faq in all_faqs:
            text = f"{faq['question']} {' '.join(faq.get('keywords', []))}"
            documents.append(f"search_document: {text}")
            metadatas.append({
                "category": faq["category"],
                "question": faq["question"],
                "answer": faq["answer"]
            })
            ids.append(faq["id"])

        # Добавляем в ChromaDB
        collection.add(documents=documents, metadatas=metadatas, ids=ids)

        logger.info(f"✅ ChromaDB переобучена: {len(all_faqs)} записей")
        
        # Уведомляем бота о необходимости перезагрузки
        notify_bot_reload()
        
        return {"success": True, "message": f"Переобучено {len(all_faqs)} записей", "count": len(all_faqs)}

    except Exception as e:
        logger.error(f"❌ Ошибка при переобучении: {e}")
        return {"success": False, "message": str(e)}


def notify_bot_reload():
    """
    Отправляет запрос всем ботам на перезагрузку коллекции
    """
    for url in ALL_BOT_RELOAD_URLS:
        try:
            response = requests.post(url, timeout=2)
            if response.status_code == 200:
                logger.info(f"✅ Бот ({url}) уведомлен о перезагрузке коллекции")
            else:
                logger.warning(f"⚠️ Бот ({url}) ответил с кодом {response.status_code}")
        except requests.exceptions.ConnectionError:
            logger.warning(f"⚠️ Не удалось связаться с ботом ({url}) (возможно, он не запущен)")
        except Exception as e:
            logger.error(f"❌ Ошибка при уведомлении бота ({url}): {e}")


def notify_bot_reload_settings():
    """
    Отправляет запрос всем ботам на перезагрузку настроек
    """
    for url in ALL_BOT_RELOAD_SETTINGS_URLS:
        try:
            response = requests.post(url, timeout=2)
            if response.status_code == 200:
                logger.info(f"✅ Бот ({url}) уведомлен о перезагрузке настроек")
            else:
                logger.warning(f"⚠️ Бот ({url}) ответил с кодом {response.status_code}")
        except requests.exceptions.ConnectionError:
            logger.warning(f"⚠️ Не удалось связаться с ботом ({url}) (возможно, он не запущен)")
        except Exception as e:
            logger.error(f"❌ Ошибка при уведомлении бота ({url}): {e}")


# ========== ЭКСПОРТ ДЛЯ АКТУАЛИЗАЦИИ ==========

def strip_bbcode(text, mark_links=False, link_marker='🔗'):
    """
    Удаляет BB коды из текста, оставляя только содержимое

    :param text: Текст с BB кодами
    :param mark_links: Если True, помечает ссылки маркером
    :param link_marker: Маркер для ссылок (по умолчанию 🔗, для PDF используйте '[ССЫЛКА]')
    :return: Чистый текст без BB кодов
    """
    if not text:
        return text

    result = text

    # Ссылки: [URL=...]текст[/URL] → маркер + текст (если mark_links=True) или просто текст
    if mark_links:
        result = re.sub(r'\[URL=([^\]]+)\](.+?)\[/URL\]', f'{link_marker} \\2', result, flags=re.IGNORECASE | re.DOTALL)
    else:
        result = re.sub(r'\[URL=([^\]]+)\](.+?)\[/URL\]', r'\2', result, flags=re.IGNORECASE | re.DOTALL)

    # Форматирование: [b]текст[/b] → текст (и аналогично для всех остальных)
    result = re.sub(r'\[b\](.+?)\[/b\]', r'\1', result, flags=re.IGNORECASE | re.DOTALL)
    result = re.sub(r'\[i\](.+?)\[/i\]', r'\1', result, flags=re.IGNORECASE | re.DOTALL)
    result = re.sub(r'\[u\](.+?)\[/u\]', r'\1', result, flags=re.IGNORECASE | re.DOTALL)
    result = re.sub(r'\[s\](.+?)\[/s\]', r'\1', result, flags=re.IGNORECASE | re.DOTALL)
    result = re.sub(r'\[code\](.+?)\[/code\]', r'\1', result, flags=re.IGNORECASE | re.DOTALL)

    return result


def convert_bbcode_to_html_safe(text):
    """
    Безопасно удаляет BB коды из текста для использования в PDF.
    Использует strip_bbcode вместо конвертации в HTML чтобы избежать ошибок парсинга.

    :param text: Текст с BB кодами
    :return: Чистый текст, безопасный для ReportLab Paragraph
    """
    if not text:
        return text

    # Удаляем BB коды, помечая ссылки текстовым маркером (т.к. PDF не поддерживает эмодзи)
    # Используем » (правая кавычка-елочка) как визуальный индикатор ссылки
    result = strip_bbcode(text, mark_links=True, link_marker='»')

    # Экранируем специальные символы XML/HTML
    result = result.replace('&', '&amp;')
    result = result.replace('<', '&lt;')
    result = result.replace('>', '&gt;')

    # Обрабатываем переносы строк
    # ReportLab Paragraph использует <br/> для переносов
    result = result.replace('\n', '<br/>')

    return result


def replace_urls_with_placeholder(text):
    """
    Заменяет все URL в тексте на короткий placeholder "[ссылка]"
    для экономии места в PDF/Excel документах.

    ВАЖНО: Эта функция должна вызываться ПОСЛЕ convert_bbcode_to_html() или strip_bbcode()

    :param text: Исходный текст с возможными URL (БЕЗ BB кодов)
    :return: Текст с замененными URL
    """
    if not text:
        return text

    # Заменяем URL на "[ссылка]"
    # Ищет: http://, https://, www., ftp:// и другие распространенные паттерны
    url_pattern = r'(?:(?:https?|ftp):\/\/)?(?:www\.)?[-a-zA-Z0-9@:%._\+~#=]{1,256}\.[a-zA-Z0-9()]{1,6}\b(?:[-a-zA-Z0-9()@:%_\+.~#?&\/=]*)'
    result = re.sub(url_pattern, '[ссылка]', text)

    return result


def generate_review_pdf(faqs, category_name):
    """
    Генерация PDF документа для актуализации FAQ
    """
    buffer = BytesIO()

    # Создаем документ (альбомная ориентация для широких таблиц)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    # Попытка использовать шрифт с поддержкой кириллицы
    font_name = 'Helvetica'  # Fallback по умолчанию

    try:
        # Список возможных путей к шрифтам с поддержкой кириллицы
        font_paths = [
            # Linux/Docker - DejaVu Sans (приоритет)
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            # Linux - альтернативные пути
            "/usr/share/fonts/truetype/dejavu-sans/DejaVuSans.ttf",
            # Windows
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/Arial.ttf",
            # Альтернативные Windows пути
            os.path.join(os.environ.get('WINDIR', 'C:/Windows'), 'Fonts', 'arial.ttf'),
        ]

        # Пробуем найти и зарегистрировать шрифт
        for font_path in font_paths:
            if os.path.exists(font_path):
                if 'DejaVu' in font_path:
                    pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
                    font_name = 'DejaVuSans'
                    logger.info(f"✓ Зарегистрирован шрифт: {font_name} ({font_path})")
                    break
                elif 'arial' in font_path.lower():
                    pdfmetrics.registerFont(TTFont('Arial', font_path))
                    font_name = 'Arial'
                    logger.info(f"✓ Зарегистрирован шрифт: {font_name} ({font_path})")
                    break

        # Если шрифт не найден, логируем предупреждение
        if font_name == 'Helvetica':
            logger.warning("⚠ Шрифт с поддержкой кириллицы не найден, используется Helvetica (кириллица не поддерживается)")
            logger.warning(f"⚠ Проверенные пути: {', '.join(font_paths)}")
    except Exception as e:
        logger.error(f"❌ Ошибка при загрузке шрифта: {e}")
        # Используем Helvetica (без кириллицы)

    # Стили
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name,
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=6
    )

    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=9,
        leading=11
    )

    # Элементы документа
    elements = []

    # Заголовок
    category_text = category_name if category_name != 'all' else 'Все категории'
    title = Paragraph(f"СПИСОК FAQ ДЛЯ АКТУАЛИЗАЦИИ", title_style)
    elements.append(title)

    # Подзаголовок с категорией и датой
    subtitle_text = f"Категория: {category_text} | Дата: {datetime.now().strftime('%d.%m.%Y')}"
    subtitle = Paragraph(subtitle_text, normal_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.5*cm))

    # Подготовка данных для таблицы
    table_data = [
        [
            Paragraph('<b>№</b>', normal_style),
            Paragraph('<b>Вопрос</b>', normal_style),
            Paragraph('<b>Ответ</b>', normal_style),
            Paragraph('<b>Ключ. слова</b>', normal_style),
            Paragraph('<b>Статус</b>', normal_style)
        ]
    ]

    for idx, faq in enumerate(faqs, 1):
        # 1. Безопасно удаляем BB коды и экранируем HTML
        question_safe = convert_bbcode_to_html_safe(faq['question'])
        answer_safe = convert_bbcode_to_html_safe(faq['answer'])

        # 2. Заменяем URL на "[ссылка]" для экономии места
        question_clean = replace_urls_with_placeholder(question_safe)
        answer_clean = replace_urls_with_placeholder(answer_safe)

        # 3. Ограничиваем длину текста для читаемости
        question_text = question_clean[:100] + '...' if len(question_clean) > 100 else question_clean
        answer_text = answer_clean[:150] + '...' if len(answer_clean) > 150 else answer_clean
        keywords_text = ', '.join(faq.get('keywords', []))[:50]

        # Статус с чекбоксами на разных строках (используем HTML br для Paragraph)
        status_text = (
            '☐ Актуально<br/>'
            '☐ Изменить<br/>'
            '☐ Удалить'
        )

        table_data.append([
            Paragraph(str(idx), normal_style),
            Paragraph(question_text, normal_style),
            Paragraph(answer_text, normal_style),
            Paragraph(keywords_text, normal_style),
            Paragraph(status_text, normal_style)
        ])

    # Создание таблицы
    col_widths = [1.5*cm, 6*cm, 8*cm, 4*cm, 4*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Стиль таблицы
    table.setStyle(TableStyle([
        # Заголовок
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        # Содержимое
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Номер по центру
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')])
    ]))

    elements.append(table)

    # Сборка документа
    try:
        doc.build(elements)
    except Exception as e:
        logger.error(f"Ошибка при генерации PDF: {e}")
        raise

    buffer.seek(0)
    return buffer


def generate_review_excel(faqs, category_name):
    """
    Генерация Excel документа для актуализации FAQ
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "FAQ для актуализации"

    # Стили
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font = Font(name='Arial', size=14, bold=True, color='1E40AF')
    subtitle_font = Font(name='Arial', size=10, color='6B7280')

    cell_alignment = Alignment(vertical='top', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Заголовок документа
    ws.merge_cells('A1:F1')
    ws['A1'] = 'СПИСОК FAQ ДЛЯ АКТУАЛИЗАЦИИ'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Подзаголовок
    category_text = category_name if category_name != 'all' else 'Все категории'
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Категория: {category_text} | Дата: {datetime.now().strftime('%d.%m.%Y')}"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Пустая строка
    ws.row_dimensions[3].height = 5

    # Заголовки столбцов (строка 4)
    headers = ['№', 'Вопрос', 'Ответ', 'Ключевые слова', 'Статус', 'Комментарий']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Высота заголовка
    ws.row_dimensions[4].height = 30

    # Ширина столбцов
    ws.column_dimensions['A'].width = 5   # №
    ws.column_dimensions['B'].width = 40  # Вопрос
    ws.column_dimensions['C'].width = 50  # Ответ
    ws.column_dimensions['D'].width = 25  # Ключевые слова
    ws.column_dimensions['E'].width = 18  # Статус
    ws.column_dimensions['F'].width = 30  # Комментарий

    # Данные FAQ
    row_num = 5
    for idx, faq in enumerate(faqs, 1):
        # Номер
        cell = ws.cell(row=row_num, column=1, value=idx)
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = border

        # Вопрос (убираем BB коды, помечая ссылки 🔗, и заменяем URL на "[ссылка]")
        question_stripped = strip_bbcode(faq['question'], mark_links=True)
        question_clean = replace_urls_with_placeholder(question_stripped)
        cell = ws.cell(row=row_num, column=2, value=question_clean)
        cell.alignment = cell_alignment
        cell.border = border

        # Ответ (убираем BB коды, помечая ссылки 🔗, и заменяем URL на "[ссылка]")
        answer_stripped = strip_bbcode(faq['answer'], mark_links=True)
        answer_clean = replace_urls_with_placeholder(answer_stripped)
        cell = ws.cell(row=row_num, column=3, value=answer_clean)
        cell.alignment = cell_alignment
        cell.border = border

        # Ключевые слова
        keywords_text = ', '.join(faq.get('keywords', []))
        cell = ws.cell(row=row_num, column=4, value=keywords_text)
        cell.alignment = cell_alignment
        cell.border = border

        # Статус (выпадающий список)
        cell = ws.cell(row=row_num, column=5, value='Актуально')  # Значение по умолчанию
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

        # Комментарий (пустая ячейка для заметок)
        cell = ws.cell(row=row_num, column=6, value='')
        cell.alignment = cell_alignment
        cell.border = border

        # Excel автоматически подберёт высоту строки при открытии файла
        # благодаря wrap_text=True в cell_alignment
        # Не устанавливаем высоту вручную - пусть Excel сделает это сам

        row_num += 1

    # Добавляем выпадающий список для колонки "Статус" (E)
    # Создаем валидацию данных
    # ВАЖНО: showDropDown=False в openpyxl означает ПОКАЗЫВАТЬ стрелку (контринтуитивно!)
    dv = DataValidation(
        type="list",
        formula1='"Актуально,Изменить,Удалить"',
        allow_blank=False,
        showDropDown=False,  # False = показывать стрелку выпадающего списка!
        showErrorMessage=True,
        errorTitle='Неверное значение',
        error='Выберите значение из списка'
    )

    # Применяем валидацию ко всем ячейкам статуса (с 5-й строки до последней)
    last_row = row_num - 1
    dv.add(f'E5:E{last_row}')
    ws.add_data_validation(dv)

    # Закрепляем первые 4 строки (заголовок + шапка таблицы)
    ws.freeze_panes = 'A5'

    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


# ========== ADMIN ROUTES ==========

@admin_bp.route('/')
def index():
    """Главная страница админки - список всех FAQ"""
    categories = database.get_all_categories()
    return render_template('admin/index.html', categories=categories)


@admin_bp.route('/faq/list')
def list_faqs():
    """Получить список FAQ (опционально по категории)"""
    category = request.args.get('category')
    if category:
        faqs = database.get_faqs_by_category(category)
    else:
        faqs = database.get_all_faqs()
    return jsonify(faqs)


@admin_bp.route('/faq/<faq_id>')
def get_faq(faq_id):
    """Получить конкретный FAQ"""
    faq = database.get_faq_by_id(faq_id)
    if faq:
        return jsonify(faq)
    return jsonify({"error": "FAQ не найден"}), 404


@admin_bp.route('/faq/add', methods=['POST'])
def add_faq():
    """Добавить новый FAQ"""
    data = request.json
    category = data.get('category')
    question = data.get('question')
    answer = data.get('answer')
    keywords = data.get('keywords', [])

    if not all([category, question, answer]):
        return jsonify({"success": False, "message": "Не все обязательные поля заполнены"}), 400

    faq_id = data.get('id') or f"faq_{uuid.uuid4().hex[:8]}"

    if isinstance(keywords, str):
        keywords = [k.strip() for k in keywords.split(',') if k.strip()]

    success = database.add_faq(faq_id, category, question, answer, keywords)
    if success:
        return jsonify({"success": True, "message": "FAQ добавлен"})
    return jsonify({"success": False, "message": "FAQ с таким ID уже существует"}), 400


@admin_bp.route('/faq/update/<faq_id>', methods=['PUT'])
def update_faq(faq_id):
    """Обновить существующий FAQ"""
    data = request.json
    category = data.get('category')
    question = data.get('question')
    answer = data.get('answer')
    keywords = data.get('keywords', [])

    if not all([category, question, answer]):
        return jsonify({"success": False, "message": "Не все обязательные поля заполнены"}), 400

    if isinstance(keywords, str):
        keywords = [k.strip() for k in keywords.split(',') if k.strip()]

    success = database.update_faq(faq_id, category, question, answer, keywords)
    if success:
        return jsonify({"success": True, "message": "FAQ обновлён"})
    return jsonify({"success": False, "message": "FAQ не найден"}), 404


@admin_bp.route('/faq/delete/<faq_id>', methods=['DELETE'])
def delete_faq(faq_id):
    """Удалить FAQ"""
    success = database.delete_faq(faq_id)
    if success:
        return jsonify({"success": True, "message": "FAQ удалён"})
    return jsonify({"success": False, "message": "FAQ не найден"}), 404


@admin_bp.route('/categories')
def get_categories():
    """Получить список всех категорий"""
    categories = database.get_all_categories()
    return jsonify(categories)


@admin_bp.route('/categories', methods=['POST'])
def add_category_route():
    """Добавить новую категорию"""
    data = request.get_json()
    category_name = data.get("name")

    if not category_name:
        return jsonify({"error": "Не указано имя категории"}), 400

    if database.add_category(category_name):
        return jsonify({"message": "Категория добавлена"}), 201
    else:
        return jsonify({"error": "Такая категория уже существует"}), 409


@admin_bp.route('/retrain', methods=['POST'])
def retrain():
    """Переобучить ChromaDB"""
    result = retrain_chromadb()
    if result["success"]:
        return jsonify(result)
    return jsonify(result), 500


@admin_bp.route('/export-review', methods=['GET'])
def export_for_review():
    """
    Экспорт FAQ для актуализации
    Параметры:
    - category: категория для фильтрации (по умолчанию 'all')
    - format: формат экспорта ('pdf' или 'excel', по умолчанию 'excel')
    """
    try:
        # Получаем параметры
        category = request.args.get('category', 'all')
        export_format = request.args.get('format', 'excel')

        # Получаем FAQ
        if category and category != 'all':
            faqs = database.get_faqs_by_category(category)
        else:
            faqs = database.get_all_faqs()

        if not faqs:
            return jsonify({
                "success": False,
                "message": "Нет данных для экспорта"
            }), 404

        # Генерируем файл
        date_str = datetime.now().strftime("%Y%m%d")

        if export_format == 'pdf':
            buffer = generate_review_pdf(faqs, category)
            mimetype = 'application/pdf'
            extension = 'pdf'
        else:  # excel
            buffer = generate_review_excel(faqs, category)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            extension = 'xlsx'

        # Формируем имя файла (для логов)
        filename_display = f'faq_review_{category}_{date_str}.{extension}'

        # Формируем безопасное имя для HTTP заголовка (ASCII-only fallback)
        filename_ascii = f'faq_review_{date_str}.{extension}'

        # URL-кодируем полное имя с кириллицей для современных браузеров (RFC 5987)
        filename_encoded = quote(filename_display.encode('utf-8'))

        # Отправляем файл
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = mimetype
        # Используем оба формата для совместимости с разными браузерами
        response.headers['Content-Disposition'] = (
            f"attachment; filename={filename_ascii}; filename*=UTF-8''{filename_encoded}"
        )

        logger.info(f"✅ Экспорт выполнен: {filename_display} ({len(faqs)} FAQ)")
        return response

    except Exception as e:
        logger.error(f"❌ Ошибка при экспорте: {e}")
        return jsonify({
            "success": False,
            "message": f"Ошибка при экспорте: {str(e)}"
        }), 500


@admin_bp.route('/search', methods=['GET'])
def search_faqs():
    """
    Поиск FAQ по тексту (в вопросах, ответах и ключевых словах)
    Параметры: ?q=текст_поиска&category=категория (опционально)
    """
    query = request.args.get('q', '').strip().lower()
    category = request.args.get('category')
    
    if not query:
        return jsonify({"success": False, "message": "Не указан поисковый запрос"}), 400
    
    try:
        # Получаем все FAQ или по категории
        if category:
            all_faqs = database.get_faqs_by_category(category)
        else:
            all_faqs = database.get_all_faqs()
        
        # Фильтруем по поисковому запросу
        results = []
        for faq in all_faqs:
            # Ищем в вопросе, ответе и ключевых словах
            question_lower = faq['question'].lower()
            answer_lower = faq['answer'].lower()
            keywords_lower = ' '.join(faq.get('keywords', [])).lower()
            
            # Проверяем совпадение
            if (query in question_lower or 
                query in answer_lower or 
                query in keywords_lower):
                
                # Добавляем информацию о том, где найдено
                match_info = []
                if query in question_lower:
                    match_info.append('вопросе')
                if query in answer_lower:
                    match_info.append('ответе')
                if query in keywords_lower:
                    match_info.append('ключевых словах')
                
                faq_copy = faq.copy()
                faq_copy['match_location'] = match_info
                results.append(faq_copy)
        
        return jsonify({
            "success": True,
            "query": query,
            "count": len(results),
            "results": results
        })
        
    except Exception as e:
        logger.error(f"Ошибка при поиске: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/search/semantic', methods=['POST'])
def semantic_search():
    """
    Семантический поиск через ChromaDB
    Body: {"query": "текст запроса", "n_results": 5}
    """
    data = request.json
    query = data.get('query', '').strip()
    n_results = data.get('n_results', 5)
    
    if not query:
        return jsonify({"success": False, "message": "Не указан поисковый запрос"}), 400
    
    try:
        # Получаем коллекцию
        try:
            collection = chroma_client.get_collection(name="faq_collection")
        except Exception:
            return jsonify({
                "success": False, 
                "message": "База знаний не инициализирована. Выполните переобучение."
            }), 404
        
        # Выполняем семантический поиск
        results = collection.query(
            query_texts=[f"search_query: {query}"],
            n_results=n_results,
            include=["documents", "metadatas", "distances"]
        )
        
        if not results or not results["documents"] or not results["documents"][0]:
            return jsonify({
                "success": True,
                "query": query,
                "count": 0,
                "results": []
            })
        
        # Формируем результаты
        search_results = []
        for i, metadata in enumerate(results["metadatas"][0]):
            distance = results["distances"][0][i]
            similarity = max(0.0, 1.0 - distance) * 100.0
            faq_id = results["ids"][0][i] if "ids" in results and results["ids"] else None
            
            search_results.append({
                "id": faq_id,
                "question": metadata["question"],
                "answer": metadata["answer"],
                "category": metadata["category"],
                "similarity": round(similarity, 1),
                "distance": round(distance, 4)
            })
        
        return jsonify({
            "success": True,
            "query": query,
            "count": len(search_results),
            "results": search_results
        })
        
    except Exception as e:
        logger.error(f"Ошибка при семантическом поиске: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ========== НАСТРОЙКИ БОТА ==========

@admin_bp.route('/settings')
def settings_page():
    """Страница настроек бота"""
    return render_template('admin/settings.html')


@admin_bp.route('/api/settings', methods=['GET'])
def get_settings():
    """Получить текущие настройки бота"""
    try:
        settings = database.get_bot_settings()
        return jsonify({
            "success": True,
            "settings": settings
        })
    except Exception as e:
        logger.error(f"Ошибка при получении настроек: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/settings', methods=['POST'])
def save_settings():
    """Сохранить настройки бота"""
    try:
        data = request.json
        settings = data.get('settings', {})

        if not settings:
            return jsonify({"success": False, "message": "Настройки не переданы"}), 400

        # Сохраняем настройки в БД
        success = database.update_bot_settings(settings)

        if success:
            # Уведомляем бота о перезагрузке настроек
            notify_bot_reload_settings()

            return jsonify({
                "success": True,
                "message": "Настройки сохранены"
            })
        else:
            return jsonify({
                "success": False,
                "message": "Ошибка при сохранении настроек"
            }), 500

    except Exception as e:
        logger.error(f"Ошибка при сохранении настроек: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/settings/reset', methods=['POST'])
def reset_settings():
    """Сбросить настройки бота к значениям по умолчанию"""
    try:
        success = database.reset_bot_settings()

        if success:
            # Уведомляем бота о перезагрузке настроек
            notify_bot_reload_settings()

            return jsonify({
                "success": True,
                "message": "Настройки сброшены к значениям по умолчанию"
            })
        else:
            return jsonify({
                "success": False,
                "message": "Ошибка при сбросе настроек"
            }), 500

    except Exception as e:
        logger.error(f"Ошибка при сбросе настроек: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/optimize-keywords', methods=['POST'])
def optimize_keywords():
    """
    Оптимизация ключевых слов через лемматизацию

    Принимает строку ключевых слов через запятую,
    лемматизирует каждое слово и возвращает уникальные леммы.

    Body: {"keywords": "претензию, претензии, товар, товары"}
    Returns: {"success": true, "optimized": "претензия, товар"}
    """
    try:
        from src.core.search import lemmatize_word

        data = request.json
        keywords_str = data.get('keywords', '')

        if not keywords_str or not keywords_str.strip():
            return jsonify({
                "success": True,
                "optimized": "",
                "original_count": 0,
                "optimized_count": 0
            })

        # Разбиваем на отдельные слова
        keywords = [kw.strip() for kw in keywords_str.split(',') if kw.strip()]

        if not keywords:
            return jsonify({
                "success": True,
                "optimized": "",
                "original_count": 0,
                "optimized_count": 0
            })

        # Лемматизируем каждое слово
        lemmatized = []
        seen = set()

        for keyword in keywords:
            lemma = lemmatize_word(keyword)

            # Добавляем только уникальные леммы
            if lemma not in seen and lemma:
                seen.add(lemma)
                lemmatized.append(lemma)

        # Собираем обратно в строку
        optimized_str = ', '.join(lemmatized)

        logger.info(f"Оптимизация ключевых слов: {len(keywords)} → {len(lemmatized)} слов")

        return jsonify({
            "success": True,
            "optimized": optimized_str,
            "original_count": len(keywords),
            "optimized_count": len(lemmatized)
        })

    except Exception as e:
        logger.error(f"Ошибка при оптимизации ключевых слов: {e}", exc_info=True)
        return jsonify({
            "success": False,
            "message": f"Ошибка оптимизации: {str(e)}"
        }), 500


# ========== ПРАВА ДОСТУПА BITRIX24 ==========

@admin_bp.route('/permissions')
def permissions_page():
    """Страница управления правами доступа Bitrix24"""
    # Получаем домен из .env
    domain = os.getenv('BITRIX24_DOMAIN', 'your-company.bitrix24.ru')
    return render_template('admin/permissions.html', domain=domain)


# ========== ЛОГИРОВАНИЕ ==========

@admin_bp.route('/logs')
def logs_page():
    """Страница просмотра логов"""
    categories = database.get_all_categories()
    return render_template('admin/logs.html', categories=categories)


@admin_bp.route('/api/logs/list', methods=['GET'])
def get_logs():
    """
    Получить список логов с фильтрацией и пагинацией
    Параметры:
    - page: номер страницы (по умолчанию 1)
    - per_page: количество записей на странице (по умолчанию 50)
    - user_id: фильтр по ID пользователя
    - faq_id: фильтр по ID FAQ
    - rating: фильтр по оценке (helpful, not_helpful, no_rating)
    - date_from: начальная дата (ISO format)
    - date_to: конечная дата (ISO format)
    - search: поиск по тексту запроса
    - no_answer: показывать только запросы без ответа (true/false)
    - platform: фильтр по платформе (telegram, bitrix24)
    """
    try:
        # Параметры пагинации
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        offset = (page - 1) * per_page

        # Параметры фильтрации
        user_id = request.args.get('user_id')
        if user_id:
            user_id = int(user_id)

        faq_id = request.args.get('faq_id')
        rating = request.args.get('rating')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        search_text = request.args.get('search')
        no_answer = request.args.get('no_answer', 'false').lower() == 'true'
        platform = request.args.get('platform')
        show_archived = request.args.get('show_archived', 'false').lower() == 'true'

        # Получаем логи
        logs, total = database.get_logs(
            limit=per_page,
            offset=offset,
            user_id=user_id,
            faq_id=faq_id,
            rating_filter=rating,
            date_from=date_from,
            date_to=date_to,
            search_text=search_text,
            no_answer=no_answer,
            platform=platform,
            show_archived=show_archived
        )

        # Вычисляем метаданные пагинации
        total_pages = (total + per_page - 1) // per_page

        return jsonify({
            "success": True,
            "logs": logs,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages
            }
        })

    except Exception as e:
        logger.error(f"Ошибка при получении логов: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/logs/statistics', methods=['GET'])
def get_logs_statistics():
    """Получить статистику по логам"""
    try:
        stats = database.get_statistics()
        # Добавляем текущий порог схожести
        stats["similarity_threshold"] = database.SIMILARITY_THRESHOLD
        return jsonify({
            "success": True,
            "statistics": stats
        })
    except Exception as e:
        logger.error(f"Ошибка при получении статистики: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/logs/rag-statistics', methods=['GET'])
def get_rag_statistics():
    """
    Получение статистики RAG генерации

    Returns:
        {
            "total_rag_answers": 150,
            "avg_tokens_per_answer": 245.3,
            "total_tokens_used": 36795,
            "rag_errors": 3,
            "rag_success_rate": 98.0,
            "models_used": {
                "openai/gpt-4o-mini": 145,
                "openai/gpt-4o": 5
            },
            "avg_chunks_per_query": 2.8,
            "avg_generation_time_ms": 1250
        }
    """
    try:
        with database.get_db_connection() as conn:
            cursor = conn.cursor()

            # Основная статистика
            cursor.execute("""
                SELECT
                    COUNT(*) as total_rag,
                    AVG(tokens_total) as avg_tokens,
                    SUM(tokens_total) as total_tokens,
                    AVG(chunks_used) as avg_chunks,
                    AVG(generation_time_ms) as avg_gen_time,
                    COUNT(CASE WHEN error_message IS NOT NULL THEN 1 END) as errors
                FROM llm_generations
                WHERE answer_log_id IN (
                    SELECT id FROM answer_logs WHERE period_id IS NULL
                )
            """)

            stats = cursor.fetchone()

            # Статистика по моделям
            cursor.execute("""
                SELECT model, COUNT(*) as count
                FROM llm_generations
                WHERE answer_log_id IN (
                    SELECT id FROM answer_logs WHERE period_id IS NULL
                )
                GROUP BY model
            """)

            models = {row['model']: row['count'] for row in cursor.fetchall()}

            total_rag = stats['total_rag'] or 0
            errors = stats['errors'] or 0

            return jsonify({
                'total_rag_answers': total_rag,
                'avg_tokens_per_answer': round(stats['avg_tokens'] or 0, 1),
                'total_tokens_used': stats['total_tokens'] or 0,
                'rag_errors': errors,
                'rag_success_rate': round((total_rag - errors) / total_rag * 100, 1) if total_rag > 0 else 0,
                'models_used': models,
                'avg_chunks_per_query': round(stats['avg_chunks'] or 0, 1),
                'avg_generation_time_ms': round(stats['avg_gen_time'] or 0)
            })

    except Exception as e:
        logger.error(f"Ошибка получения RAG статистики: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/logs/export', methods=['GET'])
def export_logs():
    """
    Экспорт логов в CSV
    Параметры: такие же как в /api/logs/list
    """
    try:

        # Параметры фильтрации (те же что и для get_logs)
        user_id = request.args.get('user_id')
        if user_id:
            user_id = int(user_id)

        faq_id = request.args.get('faq_id')
        rating = request.args.get('rating')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        search_text = request.args.get('search')
        no_answer = request.args.get('no_answer', 'false').lower() == 'true'

        # Получаем все логи
        logs, total = database.get_logs(
            limit=10000,
            offset=0,
            user_id=user_id,
            faq_id=faq_id,
            rating_filter=rating,
            date_from=date_from,
            date_to=date_to,
            search_text=search_text,
            no_answer=no_answer
        )

        # Создаем CSV в памяти
        output = BytesIO()
        wrapper = TextIOWrapper(output, encoding='utf-8-sig', newline='')

        writer = csv.writer(wrapper, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        # Заголовки
        writer.writerow([
            'Дата/Время запроса',
            'ID пользователя',
            'Имя пользователя',
            'Текст запроса',
            'Категория FAQ',
            'Вопрос FAQ',
            'Оценка схожести (%)',
            'Рейтинг',
            'Дата/Время рейтинга'
        ])

        # Данные
        for log in logs:
            # Время уже конвертировано в UTC+7 функцией database.get_logs()
            query_timestamp = log.get('query_timestamp', '')
            if query_timestamp:
                query_timestamp = query_timestamp + ' UTC+7'

            rating_timestamp = log.get('rating_timestamp', '')
            if rating_timestamp:
                rating_timestamp = rating_timestamp + ' UTC+7'

            user_id_val = log.get('user_id')
            similarity = round(log.get('similarity_score', 0), 1) if log.get('similarity_score') is not None else ''
            rating_val = log.get('rating', '')

            writer.writerow([
                query_timestamp,
                int(user_id_val) if user_id_val is not None else '',
                log.get('username', ''),
                log.get('query_text', ''),
                log.get('category', ''),
                log.get('faq_question', ''),
                similarity,
                rating_val,
                rating_timestamp
            ])

        # Flush TextIOWrapper, чтобы данные попали в BytesIO
        wrapper.flush()

        # Теперь можно безопасно получить байты
        resp = make_response(output.getvalue())
        resp.headers["Content-Disposition"] = "attachment; filename=logs_export.csv"
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"

        return resp

    except Exception as e:
        logger.error(f"Ошибка при экспорте логов: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ========== ТЕСТОВЫЕ ПЕРИОДЫ ==========

@admin_bp.route('/test-periods')
def test_periods_page():
    """Страница управления тестовыми периодами"""
    return render_template('admin/test_periods.html')


@admin_bp.route('/api/test-periods/list', methods=['GET'])
def get_test_periods():
    """Получить список всех тестовых периодов"""
    try:
        periods = database.get_test_periods()
        return jsonify({
            "success": True,
            "periods": periods
        })
    except Exception as e:
        logger.error(f"Ошибка при получении тестовых периодов: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/test-periods/active', methods=['GET'])
def get_active_test_period():
    """Получить активный тестовый период"""
    try:
        period = database.get_active_test_period()
        return jsonify({
            "success": True,
            "period": period
        })
    except Exception as e:
        logger.error(f"Ошибка при получении активного периода: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/test-periods/create', methods=['POST'])
def create_test_period():
    """Создать новый тестовый период"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()

        if not name:
            return jsonify({"success": False, "message": "Название периода обязательно"}), 400

        period_id = database.create_test_period(name, description)

        if period_id:
            return jsonify({
                "success": True,
                "message": f"Тестовый период '{name}' создан",
                "period_id": period_id
            })
        else:
            return jsonify({
                "success": False,
                "message": "Не удалось создать период. Возможно, уже есть активный период."
            }), 400

    except Exception as e:
        logger.error(f"Ошибка при создании тестового периода: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/test-periods/<int:period_id>/end', methods=['POST'])
def end_test_period(period_id):
    """Завершить тестовый период"""
    try:
        success = database.end_test_period(period_id)

        if success:
            return jsonify({
                "success": True,
                "message": "Тестовый период завершён"
            })
        else:
            return jsonify({
                "success": False,
                "message": "Не удалось завершить период"
            }), 400

    except Exception as e:
        logger.error(f"Ошибка при завершении тестового периода: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/test-periods/<int:period_id>/archive', methods=['POST'])
def archive_period_logs(period_id):
    """Архивировать текущие логи в тестовый период"""
    try:
        result = database.archive_current_logs(period_id)

        return jsonify({
            "success": True,
            "message": f"Заархивировано: {result['queries']} запросов, {result['answers']} ответов, {result['ratings']} оценок",
            "archived": result
        })

    except Exception as e:
        logger.error(f"Ошибка при архивации логов: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/test-periods/clear-unarchived', methods=['POST'])
def clear_unarchived_logs():
    """Удалить неархивированные логи"""
    try:
        result = database.clear_unarchived_logs()

        return jsonify({
            "success": True,
            "message": f"Удалено: {result['queries']} запросов, {result['answers']} ответов, {result['ratings']} оценок",
            "deleted": result
        })

    except Exception as e:
        logger.error(f"Ошибка при очистке логов: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/test-periods/<int:period_id>/statistics', methods=['GET'])
def get_period_statistics(period_id):
    """Получить статистику по тестовому периоду"""
    try:
        stats = database.get_period_statistics(period_id)

        if not stats:
            return jsonify({
                "success": False,
                "message": "Тестовый период не найден"
            }), 404

        return jsonify({
            "success": True,
            "statistics": stats
        })

    except Exception as e:
        logger.error(f"Ошибка при получении статистики периода: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/test-periods/<int:period_id>/export', methods=['GET'])
def export_period_report(period_id):
    """Экспорт отчета по тестовому периоду"""
    try:
        from src.web.report_generator import (
            generate_period_excel_report,
            generate_period_json_report,
            generate_period_csv_report
        )

        # Получаем статистику
        stats = database.get_period_statistics(period_id)

        if not stats:
            return jsonify({
                "success": False,
                "message": "Тестовый период не найден"
            }), 404

        # Определяем формат экспорта
        export_format = request.args.get('format', 'excel').lower()

        if export_format == 'excel':
            buffer = generate_period_excel_report(stats)
            filename = f"test_period_{period_id}_{stats['period']['name']}.xlsx"

            resp = make_response(buffer.getvalue())
            resp.headers["Content-Disposition"] = f"attachment; filename={quote(filename)}"
            resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return resp

        elif export_format == 'json':
            json_data = generate_period_json_report(stats)
            filename = f"test_period_{period_id}.json"

            resp = make_response(json_data)
            resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
            resp.headers["Content-Type"] = "application/json; charset=utf-8"
            return resp

        elif export_format == 'csv':
            buffer = generate_period_csv_report(stats)
            filename = f"test_period_{period_id}.csv"

            resp = make_response(buffer.getvalue())
            resp.headers["Content-Disposition"] = f"attachment; filename={quote(filename)}"
            resp.headers["Content-Type"] = "text/csv; charset=utf-8"
            return resp

        else:
            return jsonify({
                "success": False,
                "message": f"Неподдерживаемый формат: {export_format}"
            }), 400

    except Exception as e:
        logger.error(f"Ошибка при экспорте отчета: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route('/api/test-periods/<int:period_id>/failed-queries', methods=['GET'])
def get_period_failed_queries(period_id):
    """Получить неудачные запросы для работы над ошибками"""
    try:
        limit = int(request.args.get('limit', 100))
        failed_queries = database.get_failed_queries_for_period(period_id, limit)

        return jsonify({
            "success": True,
            "failed_queries": failed_queries
        })

    except Exception as e:
        logger.error(f"Ошибка при получении неудачных запросов: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ========== PUBLIC ROUTES (временные заглушки) ==========

@app.route('/')
def public_search():
    """
    Корневой роут - в production закрыт, в dev показывает публичный поиск
    """
    if is_production():
        return jsonify({
            'error': 'Доступ запрещен',
            'message': 'Доступ возможен только через Битрикс24',
            'redirect': os.getenv('BITRIX24_DOMAIN', '')
        }), 403

    # В dev режиме показываем публичный поиск
    return render_template('search.html')


@app.route('/api/search', methods=['POST'])
def public_api_search():
    """
    API для публичного семантического поиска
    В production режиме закрыт
    """
    if is_production():
        return jsonify({
            'error': 'Доступ запрещен',
            'message': 'Публичный API недоступен в production режиме'
        }), 403

    data = request.json
    query = data.get('query', '').strip()
    user_id = data.get('user_id', 0)  # Для веба используем 0 или сессионный ID

    if not query:
        return jsonify({"success": False, "message": "Не указан поисковый запрос"}), 400

    try:
        # Логируем запрос пользователя
        query_log_id = database.add_query_log(
            user_id=user_id,
            username='web_user',
            query_text=query,
            platform='web'
        )

        # Получаем коллекцию
        try:
            collection = chroma_client.get_collection(name="faq_collection")
        except Exception:
            return jsonify({
                "success": False,
                "message": "База знаний не инициализирована."
            }), 404

        # Выполняем семантический поиск
        results = collection.query(
            query_texts=[f"search_query: {query}"],
            n_results=5,
            include=["documents", "metadatas", "distances"]
        )

        if not results or not results["documents"] or not results["documents"][0]:
            return jsonify({
                "success": True,
                "query": query,
                "count": 0,
                "results": []
            })

        # Формируем результаты
        search_results = []
        for i, metadata in enumerate(results["metadatas"][0]):
            distance = results["distances"][0][i]
            similarity = max(0.0, 1.0 - distance) * 100.0
            faq_id = results["ids"][0][i] if "ids" in results and results["ids"] else None

            # Применяем порог схожести
            if similarity >= database.SIMILARITY_THRESHOLD:
                # Логируем показанный ответ
                answer_log_id = database.add_answer_log(
                    query_log_id=query_log_id,
                    faq_id=faq_id,
                    similarity_score=similarity,
                    answer_shown=metadata["answer"]
                )

                search_results.append({
                    "id": faq_id,
                    "answer_log_id": answer_log_id,  # Добавляем для обратной связи
                    "question": metadata["question"],
                    "answer": metadata["answer"],
                    "category": metadata["category"],
                    "similarity": round(similarity, 1)
                })

        return jsonify({
            "success": True,
            "query": query,
            "count": len(search_results),
            "results": search_results
        })

    except Exception as e:
        logger.error(f"Ошибка при публичном поиске: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/feedback', methods=['POST'])
def public_feedback():
    """API для сохранения обратной связи от пользователей"""
    data = request.json
    answer_log_id = data.get('answer_log_id')
    rating = data.get('rating')  # 'helpful' или 'not_helpful'
    user_id = data.get('user_id', 0)  # Для веб-версии можем использовать 0 или генерировать

    if not answer_log_id or not rating:
        return jsonify({"success": False, "message": "Не все поля заполнены"}), 400

    try:
        database.add_rating_log(answer_log_id, user_id, rating)
        return jsonify({"success": True, "message": "Спасибо за обратную связь!"})
    except Exception as e:
        logger.error(f"Ошибка при сохранении обратной связи: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/admin/api/search-level-stats', methods=['GET'])
def api_search_level_stats():
    """API: Статистика по уровням каскадного поиска"""
    try:
        stats = database.get_search_level_statistics()
        return jsonify(stats)
    except Exception as e:
        logger.error(f"Ошибка получения статистики уровней поиска: {e}")
        return jsonify({'error': str(e)}), 500


# ========== BITRIX24 INTEGRATION ==========

@app.route('/bitrix24/install', methods=['GET', 'POST'])
def bitrix24_install():
    """Обработчик установки приложения Битрикс24"""
    return handle_install(request)


@app.route('/bitrix24/index', methods=['GET', 'POST'])
def bitrix24_index():
    """Обработчик первого открытия приложения Битрикс24"""
    return handle_index(request)


@app.route('/bitrix24/app', methods=['GET', 'POST'])
def bitrix24_app():
    """Страница встраиваемого приложения Битрикс24"""
    return handle_app(request)


# Health check endpoint для Docker healthcheck
@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint для мониторинга"""
    try:
        # Проверяем доступность базы данных
        with database.get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM faq")
            faq_count = cursor.fetchone()[0]

        # Проверяем ChromaDB
        chromadb_count = 0
        try:
            collection = chroma_client.get_collection(name="faq_collection")
            chromadb_count = collection.count()
        except Exception:
            # Коллекция ещё не создана (до первого переобучения)
            chromadb_count = 0

        return jsonify({
            'status': 'ok',
            'database': 'connected',
            'faq_count': faq_count,
            'chromadb_records': chromadb_count
        }), 200
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 503


# Регистрируем Blueprint для управления правами Битрикс24
app.register_blueprint(bitrix24_permissions_bp, url_prefix='/api/bitrix24/permissions')

# Регистрируем Blueprint админки
app.register_blueprint(admin_bp)


# ========== MAIN ==========

if __name__ == '__main__':
    database.init_database()
    print("🌐 Веб-интерфейс запущен на http://127.0.0.1:5000")
    print("📝 Используйте этот интерфейс для управления FAQ")
    app.run(debug=False, host='0.0.0.0', port=5000)