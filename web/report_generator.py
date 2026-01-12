# -*- coding: utf-8 -*-
"""
Модуль для генерации отчетов по тестовым периодам
Поддержка: Excel с графиками, JSON, CSV
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from io import BytesIO, TextIOWrapper
import json
import csv
from datetime import datetime


def generate_period_excel_report(stats: dict) -> BytesIO:
    """
    Генерация Excel отчета по тестовому периоду с графиками

    :param stats: Статистика по периоду из get_period_statistics()
    :return: BytesIO объект с Excel файлом
    """
    wb = Workbook()

    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=11)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== ЛИСТ 1: ОБЩАЯ СВОДКА ==========
    ws_summary = wb.active
    ws_summary.title = "📊 Сводка"

    # Заголовок отчета
    ws_summary['A1'] = f"Отчет по тестовому периоду: {stats['period']['name']}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')

    # Информация о периоде
    row = 3
    ws_summary[f'A{row}'] = "Период:"
    ws_summary[f'B{row}'] = f"{stats['period']['start_date']} — {stats['period']['end_date'] or 'Активен'}"
    ws_summary[f'A{row}'].font = Font(bold=True)

    row += 1
    if stats['period'].get('description'):
        ws_summary[f'A{row}'] = "Описание:"
        ws_summary[f'B{row}'] = stats['period']['description']
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

    row += 1

    # Ключевые метрики
    ws_summary[f'A{row}'] = "КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ"
    ws_summary[f'A{row}'].font = Font(bold=True, size=13, color="4472C4")
    row += 1

    metrics = [
        ("Всего запросов", stats['total_queries']),
        ("Уникальных пользователей", stats['unique_users']),
        ("Всего ответов", stats['total_answers']),
        ("Запросов без ответа", stats['no_answer_count']),
        ("Средняя уверенность", f"{stats['avg_similarity']}%"),
        ("", ""),
        ("Положительных оценок", stats['helpful_count']),
        ("Отрицательных оценок", stats['not_helpful_count']),
        ("% полезности", f"{stats['helpful_percentage']}%"),
    ]

    for label, value in metrics:
        if label:
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

    # Автоширина колонок
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20

    # ========== ЛИСТ 2: РАСПРЕДЕЛЕНИЕ ПО УРОВНЯМ ПОИСКА ==========
    ws_levels = wb.create_sheet(title="🔍 Уровни поиска")

    # Заголовок
    ws_levels['A1'] = "Распределение по уровням каскадного поиска"
    ws_levels['A1'].font = Font(bold=True, size=13)
    ws_levels.merge_cells('A1:D1')

    # Таблица
    row = 3
    headers = ['Уровень', 'Иконка', 'Количество', 'Средняя уверенность (%)']
    for col, header in enumerate(headers, start=1):
        cell = ws_levels.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Иконки и названия уровней
    level_names = {
        'exact': ('Точное совпадение', '🎯'),
        'keyword': ('Поиск по ключевым словам', '🔑'),
        'semantic': ('Семантический поиск', '🧠'),
        'disambiguation_shown': ('Показ вариантов для выбора', '🔀'),
        'disambiguation': ('Выбор из вариантов', '✅'),
        'direct': ('Прямой выбор', '📄'),
        'none': ('Не найдено', '❌')
    }

    row += 1
    for level_key in ['exact', 'keyword', 'semantic', 'disambiguation_shown', 'disambiguation', 'direct', 'none']:
        level_data = stats['search_levels'].get(level_key, {'count': 0, 'avg_confidence': 0})
        level_name, icon = level_names.get(level_key, (level_key, ''))

        ws_levels.cell(row=row, column=1).value = level_name
        ws_levels.cell(row=row, column=2).value = icon
        ws_levels.cell(row=row, column=3).value = level_data['count']
        ws_levels.cell(row=row, column=4).value = level_data['avg_confidence']

        for col in range(1, 5):
            ws_levels.cell(row=row, column=col).border = border
        row += 1

    # График - круговая диаграмма
    chart = PieChart()
    chart.title = "Распределение запросов по уровням поиска"
    chart.height = 12
    chart.width = 18

    # Данные для графика
    data = Reference(ws_levels, min_col=3, min_row=3, max_row=row-1)
    labels = Reference(ws_levels, min_col=1, min_row=4, max_row=row-1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    ws_levels.add_chart(chart, "F3")

    # Автоширина
    ws_levels.column_dimensions['A'].width = 30
    ws_levels.column_dimensions['B'].width = 10
    ws_levels.column_dimensions['C'].width = 15
    ws_levels.column_dimensions['D'].width = 25

    # ========== ЛИСТ 3: ПОПУЛЯРНЫЕ ЗАПРОСЫ ==========
    ws_top = wb.create_sheet(title="⭐ Популярные запросы")

    ws_top['A1'] = f"Топ-{len(stats['top_queries'])} популярных запросов"
    ws_top['A1'].font = Font(bold=True, size=13)
    ws_top.merge_cells('A1:C1')

    # Таблица
    row = 3
    headers = ['№', 'Запрос', 'Количество']
    for col, header in enumerate(headers, start=1):
        cell = ws_top.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row += 1
    for idx, query_data in enumerate(stats['top_queries'], start=1):
        ws_top.cell(row=row, column=1).value = idx
        ws_top.cell(row=row, column=2).value = query_data['query']
        ws_top.cell(row=row, column=3).value = query_data['count']

        for col in range(1, 4):
            ws_top.cell(row=row, column=col).border = border
        row += 1

    # График
    if stats['top_queries']:
        chart = BarChart()
        chart.title = "Частота запросов"
        chart.x_axis.title = "Количество"
        chart.y_axis.title = "Запрос"
        chart.height = 12
        chart.width = 18

        data = Reference(ws_top, min_col=3, min_row=3, max_row=row-1)
        categories = Reference(ws_top, min_col=2, min_row=4, max_row=row-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws_top.add_chart(chart, "E3")

    ws_top.column_dimensions['A'].width = 5
    ws_top.column_dimensions['B'].width = 60
    ws_top.column_dimensions['C'].width = 15

    # ========== ЛИСТ 4: ЛУЧШИЕ FAQ ==========
    ws_helpful = wb.create_sheet(title="👍 Лучшие FAQ")

    ws_helpful['A1'] = "FAQ с наилучшими оценками"
    ws_helpful['A1'].font = Font(bold=True, size=13)
    ws_helpful.merge_cells('A1:D1')

    row = 3
    headers = ['№', 'Вопрос', 'Категория', 'Положительных оценок']
    for col, header in enumerate(headers, start=1):
        cell = ws_helpful.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row += 1
    for idx, faq in enumerate(stats['top_helpful_faqs'], start=1):
        ws_helpful.cell(row=row, column=1).value = idx
        ws_helpful.cell(row=row, column=2).value = faq['question']
        ws_helpful.cell(row=row, column=3).value = faq['category']
        ws_helpful.cell(row=row, column=4).value = faq['helpful_count']

        for col in range(1, 5):
            ws_helpful.cell(row=row, column=col).border = border
        row += 1

    ws_helpful.column_dimensions['A'].width = 5
    ws_helpful.column_dimensions['B'].width = 50
    ws_helpful.column_dimensions['C'].width = 20
    ws_helpful.column_dimensions['D'].width = 20

    # ========== ЛИСТ 5: ТРЕБУЮТ УЛУЧШЕНИЯ ==========
    ws_improve = wb.create_sheet(title="⚠️ Требуют улучшения")

    ws_improve['A1'] = "FAQ с низкими оценками (работа над ошибками)"
    ws_improve['A1'].font = Font(bold=True, size=13, color="C00000")
    ws_improve.merge_cells('A1:D1')

    row = 3
    headers = ['№', 'Вопрос', 'Категория', 'Отрицательных оценок']
    for col, header in enumerate(headers, start=1):
        cell = ws_improve.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = border

    row += 1
    for idx, faq in enumerate(stats['need_improvement_faqs'], start=1):
        ws_improve.cell(row=row, column=1).value = idx
        ws_improve.cell(row=row, column=2).value = faq['question']
        ws_improve.cell(row=row, column=3).value = faq['category']
        ws_improve.cell(row=row, column=4).value = faq['not_helpful_count']

        for col in range(1, 5):
            ws_improve.cell(row=row, column=col).border = border
        row += 1

    ws_improve.column_dimensions['A'].width = 5
    ws_improve.column_dimensions['B'].width = 50
    ws_improve.column_dimensions['C'].width = 20
    ws_improve.column_dimensions['D'].width = 20

    # ========== ЛИСТ 6: ДИНАМИКА ПО ДНЯМ ==========
    if stats['daily_dynamics']:
        ws_daily = wb.create_sheet(title="📅 Динамика")

        ws_daily['A1'] = "Динамика запросов по дням"
        ws_daily['A1'].font = Font(bold=True, size=13)
        ws_daily.merge_cells('A1:C1')

        row = 3
        headers = ['Дата', 'Количество запросов']
        for col, header in enumerate(headers, start=1):
            cell = ws_daily.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        row += 1
        for day_data in stats['daily_dynamics']:
            ws_daily.cell(row=row, column=1).value = day_data['date']
            ws_daily.cell(row=row, column=2).value = day_data['count']

            for col in range(1, 3):
                ws_daily.cell(row=row, column=col).border = border
            row += 1

        # Линейный график
        chart = LineChart()
        chart.title = "Динамика запросов"
        chart.y_axis.title = "Количество запросов"
        chart.x_axis.title = "Дата"
        chart.height = 12
        chart.width = 18

        data = Reference(ws_daily, min_col=2, min_row=3, max_row=row-1)
        categories = Reference(ws_daily, min_col=1, min_row=4, max_row=row-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws_daily.add_chart(chart, "D3")

        ws_daily.column_dimensions['A'].width = 15
        ws_daily.column_dimensions['B'].width = 20

    # ========== ЛИСТ 7: ПЛАТФОРМЫ ==========
    if stats['platforms']:
        ws_platforms = wb.create_sheet(title="💻 Платформы")

        ws_platforms['A1'] = "Распределение по платформам"
        ws_platforms['A1'].font = Font(bold=True, size=13)
        ws_platforms.merge_cells('A1:C1')

        row = 3
        headers = ['Платформа', 'Количество запросов']
        for col, header in enumerate(headers, start=1):
            cell = ws_platforms.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        row += 1
        for platform, count in stats['platforms'].items():
            ws_platforms.cell(row=row, column=1).value = platform
            ws_platforms.cell(row=row, column=2).value = count

            for col in range(1, 3):
                ws_platforms.cell(row=row, column=col).border = border
            row += 1

        # Круговая диаграмма
        chart = PieChart()
        chart.title = "Распределение по платформам"
        chart.height = 12
        chart.width = 18

        data = Reference(ws_platforms, min_col=2, min_row=3, max_row=row-1)
        labels = Reference(ws_platforms, min_col=1, min_row=4, max_row=row-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        ws_platforms.add_chart(chart, "D3")

        ws_platforms.column_dimensions['A'].width = 20
        ws_platforms.column_dimensions['B'].width = 20

    # ========== ЛИСТ 8: НЕУДАЧНЫЕ ЗАПРОСЫ ==========
    if stats.get('failed_queries') and len(stats['failed_queries']) > 0:
        ws_failed = wb.create_sheet(title="❌ Неудачные запросы")

        ws_failed['A1'] = "Неудачные запросы (для работы над ошибками)"
        ws_failed['A1'].font = Font(bold=True, size=13, color="FF0000")
        ws_failed.merge_cells('A1:E1')

        ws_failed['A2'] = f"Всего неудачных запросов: {len(stats['failed_queries'])}"
        ws_failed['A2'].font = Font(italic=True, size=10)
        ws_failed.merge_cells('A2:E2')

        row = 4
        headers = ['Дата/Время', 'Пользователь', 'Платформа', 'Запрос', 'Причина']
        for col, header in enumerate(headers, start=1):
            cell = ws_failed.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        row += 1
        for query in stats['failed_queries']:
            # Определяем причину неудачи
            if not query['faq_id']:
                reason = "❌ Ответ не найден"
            elif query['rating'] == 'not_helpful':
                reason = f"👎 Негативная оценка (FAQ: {query['faq_question'][:50]}...)"
            elif query['similarity_score'] and query['similarity_score'] < 45:
                reason = f"⚠️ Низкая уверенность ({query['similarity_score']:.0f}%)"
            else:
                reason = "❓ Неизвестная причина"

            ws_failed.cell(row=row, column=1).value = query['timestamp']
            ws_failed.cell(row=row, column=2).value = query['username'] or 'Аноним'
            ws_failed.cell(row=row, column=3).value = query['platform']
            ws_failed.cell(row=row, column=4).value = query['query_text']
            ws_failed.cell(row=row, column=5).value = reason

            # Применяем границы и выравнивание
            for col in range(1, 6):
                cell = ws_failed.cell(row=row, column=col)
                cell.border = border
                if col == 4:  # Запрос
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            row += 1

        # Настройка ширины колонок
        ws_failed.column_dimensions['A'].width = 18
        ws_failed.column_dimensions['B'].width = 20
        ws_failed.column_dimensions['C'].width = 12
        ws_failed.column_dimensions['D'].width = 50
        ws_failed.column_dimensions['E'].width = 40

    # Сохранение в BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def generate_period_json_report(stats: dict) -> str:
    """
    Генерация JSON отчета по тестовому периоду

    :param stats: Статистика по периоду
    :return: JSON строка
    """
    return json.dumps(stats, ensure_ascii=False, indent=2)


def generate_period_csv_report(stats: dict) -> BytesIO:
    """
    Генерация CSV отчета по тестовому периоду (сводка)

    :param stats: Статистика по периоду
    :return: BytesIO объект с CSV файлом
    """
    buffer = BytesIO()
    writer = csv.writer(TextIOWrapper(buffer, encoding='utf-8-sig', newline=''))

    # Заголовок
    writer.writerow([f"Отчет по тестовому периоду: {stats['period']['name']}"])
    writer.writerow([])

    # Общая информация
    writer.writerow(['Метрика', 'Значение'])
    writer.writerow(['Период', f"{stats['period']['start_date']} — {stats['period']['end_date'] or 'Активен'}"])
    writer.writerow(['Всего запросов', stats['total_queries']])
    writer.writerow(['Уникальных пользователей', stats['unique_users']])
    writer.writerow(['Всего ответов', stats['total_answers']])
    writer.writerow(['Запросов без ответа', stats['no_answer_count']])
    writer.writerow(['Средняя уверенность (%)', stats['avg_similarity']])
    writer.writerow(['Положительных оценок', stats['helpful_count']])
    writer.writerow(['Отрицательных оценок', stats['not_helpful_count']])
    writer.writerow(['% полезности', stats['helpful_percentage']])
    writer.writerow([])

    # Уровни поиска
    writer.writerow(['Распределение по уровням поиска'])
    writer.writerow(['Уровень', 'Количество', 'Средняя уверенность (%)'])
    for level, data in stats['search_levels'].items():
        writer.writerow([level, data['count'], data['avg_confidence']])
    writer.writerow([])

    # Популярные запросы
    writer.writerow(['Топ популярных запросов'])
    writer.writerow(['№', 'Запрос', 'Количество'])
    for idx, query in enumerate(stats['top_queries'], start=1):
        writer.writerow([idx, query['query'], query['count']])
    writer.writerow([])

    # Лучшие FAQ
    writer.writerow(['FAQ с лучшими оценками'])
    writer.writerow(['№', 'Вопрос', 'Категория', 'Положительных оценок'])
    for idx, faq in enumerate(stats['top_helpful_faqs'], start=1):
        writer.writerow([idx, faq['question'], faq['category'], faq['helpful_count']])
    writer.writerow([])

    # FAQ требующие улучшения
    writer.writerow(['FAQ требующие улучшения'])
    writer.writerow(['№', 'Вопрос', 'Категория', 'Отрицательных оценок'])
    for idx, faq in enumerate(stats['need_improvement_faqs'], start=1):
        writer.writerow([idx, faq['question'], faq['category'], faq['not_helpful_count']])

    buffer.seek(0)
    return buffer
